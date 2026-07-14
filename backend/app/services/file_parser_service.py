import asyncio
import csv
import datetime as dt
import json
import math
from collections import Counter
from pathlib import Path
from typing import Any

from pypdf import PdfReader
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.uploaded_file import FileStatus, UploadedFile
from app.services.nlp_service import analyze_document_async
from app.utils.text_utils import data_preview, truncate_text


def _safe_json_value(value: Any) -> Any:
    if value is None or (isinstance(value, float) and not math.isfinite(value)):
        return None
    if isinstance(value, (dt.date, dt.datetime)):
        return value.isoformat()
    return value


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        number = float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _to_month(value: Any) -> str | None:
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m")
    if value is None:
        return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%Y-%m"):
        try:
            return dt.datetime.strptime(text[:10], fmt).strftime("%Y-%m")
        except ValueError:
            continue
    return None


def _median(values: list[float]) -> float:
    ordered = sorted(values)
    middle = len(ordered) // 2
    if len(ordered) % 2:
        return ordered[middle]
    return (ordered[middle - 1] + ordered[middle]) / 2


def _quantile(values: list[float], q: float) -> float:
    ordered = sorted(values)
    if not ordered:
        return 0.0
    position = (len(ordered) - 1) * q
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return ordered[int(position)]
    weight = position - lower
    return ordered[lower] * (1 - weight) + ordered[upper] * weight


def _correlation(first: list[float], second: list[float]) -> float | None:
    if len(first) < 2 or len(second) < 2 or len(first) != len(second):
        return None
    first_mean = sum(first) / len(first)
    second_mean = sum(second) / len(second)
    numerator = sum((a - first_mean) * (b - second_mean) for a, b in zip(first, second))
    first_denominator = math.sqrt(sum((a - first_mean) ** 2 for a in first))
    second_denominator = math.sqrt(sum((b - second_mean) ** 2 for b in second))
    denominator = first_denominator * second_denominator
    return numerator / denominator if denominator else None


def _tabular_profile(rows: list[dict[str, Any]], columns: list[str]) -> dict[str, Any]:
    missing = {
        column: count
        for column in columns
        if (count := sum(1 for row in rows if row.get(column) in (None, ""))) > 0
    }
    numeric_values: dict[str, list[float]] = {}
    for column in columns:
        values = [_to_float(row.get(column)) for row in rows if row.get(column) not in (None, "")]
        numbers = [value for value in values if value is not None]
        if numbers and len(numbers) >= max(1, int(len(values) * 0.8)):
            numeric_values[column] = numbers
    numeric_columns = list(numeric_values.keys())[:12]

    numeric_profiles: list[dict[str, Any]] = []
    anomalies: list[dict[str, Any]] = []
    for column in numeric_columns:
        series = numeric_values[column]
        if not series:
            continue
        q1, q3 = _quantile(series, 0.25), _quantile(series, 0.75)
        iqr = q3 - q1
        lower, upper = q1 - 1.5 * iqr, q3 + 1.5 * iqr
        outliers = [value for value in series if value < lower or value > upper]
        numeric_profiles.append(
            {
                "column": column,
                "count": len(series),
                "mean": round(sum(series) / len(series), 4),
                "median": round(_median(series), 4),
                "min": round(min(series), 4),
                "max": round(max(series), 4),
                "sum": round(sum(series), 4),
            }
        )
        if outliers:
            anomalies.append(
                {
                    "column": column,
                    "count": len(outliers),
                    "lower_bound": round(lower, 4),
                    "upper_bound": round(upper, 4),
                    "sample_values": [round(value, 4) for value in outliers[:5]],
                }
            )

    categorical_profiles: list[dict[str, Any]] = []
    categorical_columns = [column for column in columns if column not in numeric_columns][:10]
    for column in categorical_columns:
        counts = Counter(
            str(row.get(column) if row.get(column) not in (None, "") else "Missing")
            for row in rows
        ).most_common(8)
        categorical_profiles.append(
            {
                "column": column,
                "labels": [label for label, _ in counts],
                "values": [count for _, count in counts],
            }
        )

    time_series: list[dict[str, Any]] = []
    for column in columns:
        if len(time_series) >= 3:
            break
        if not any(term in column.lower() for term in ("date", "time", "month", "year")):
            continue
        periods = [_to_month(row.get(column)) for row in rows]
        counts = Counter(period for period in periods if period)
        if sum(counts.values()) < 2:
            continue
        grouped = sorted(counts.items())
        time_series.append(
            {
                "column": column,
                "labels": [period for period, _ in grouped],
                "values": [count for _, count in grouped],
                "metric": "record_count",
            }
        )

    correlations: list[dict[str, Any]] = []
    if len(numeric_columns) >= 2:
        pairs: list[tuple[float, float, str, str]] = []
        for index, first in enumerate(numeric_columns):
            for second in numeric_columns[index + 1 :]:
                paired = [
                    (_to_float(row.get(first)), _to_float(row.get(second)))
                    for row in rows
                ]
                valid = [(a, b) for a, b in paired if a is not None and b is not None]
                value = _correlation([a for a, _ in valid], [b for _, b in valid])
                if value is not None:
                    pairs.append((abs(value), value, first, second))
        correlations = [
            {"first": first, "second": second, "coefficient": round(value, 4)}
            for _, value, first, second in sorted(pairs, reverse=True)[:8]
        ]

    return {
        "missing_values": missing,
        "numeric_profiles": numeric_profiles,
        "categorical_profiles": categorical_profiles,
        "time_series": time_series,
        "anomalies": anomalies,
        "correlations": correlations,
    }


def _parse_tabular(path: Path, file_type: str) -> tuple[str, dict[str, Any]]:
    if file_type == "csv":
        with path.open("r", encoding="utf-8-sig", newline="", errors="replace") as handle:
            reader = csv.DictReader(handle)
            columns = [str(column) for column in (reader.fieldnames or [])]
            rows = [{str(key): _safe_json_value(value) for key, value in row.items()} for row in reader]
    else:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        raw_rows = sheet.iter_rows(values_only=True)
        headers = next(raw_rows, ())
        columns = [str(value) if value is not None else f"Column {index}" for index, value in enumerate(headers, start=1)]
        rows = [
            {
                columns[index]: _safe_json_value(value)
                for index, value in enumerate(row[: len(columns)])
            }
            for row in raw_rows
        ]
        workbook.close()

    preview = rows[:100]
    parsed = {
        "columns": columns,
        "row_count": len(rows),
        "column_count": len(columns),
        "preview": preview,
        "analytics": _tabular_profile(rows, columns),
    }
    lines = [
        f"Dataset contains {parsed['row_count']} records and {parsed['column_count']} columns.",
        f"Columns: {', '.join(parsed['columns'])}.",
    ]
    for profile in parsed["analytics"]["numeric_profiles"]:
        lines.append(
            f"{profile['column']} has an average of {profile['mean']}, median "
            f"{profile['median']}, minimum {profile['min']}, and maximum {profile['max']}."
        )
    for category in parsed["analytics"]["categorical_profiles"][:5]:
        distribution = ", ".join(
            f"{label}: {value}"
            for label, value in zip(category["labels"], category["values"])
        )
        lines.append(f"{category['column']} distribution is {distribution}.")
    for anomaly in parsed["analytics"]["anomalies"]:
        lines.append(
            f"{anomaly['column']} contains {anomaly['count']} anomalous values. "
            f"Examples: {', '.join(str(value) for value in anomaly['sample_values'])}."
        )
    lines.append("Representative records:")
    for index, row in enumerate(preview[:100], start=1):
        values = "; ".join(f"{key}: {value}" for key, value in row.items())
        lines.append(f"Record {index}: {values}.")
    return truncate_text("\n".join(lines)), parsed


def _parse_file(path: Path, file_type: str) -> tuple[str | None, dict | list | None]:
    if file_type in {"csv", "xlsx"}:
        return _parse_tabular(path, file_type)
    if file_type == "txt":
        return truncate_text(path.read_text(encoding="utf-8", errors="replace")), None
    if file_type == "json":
        data = json.loads(path.read_text(encoding="utf-8"))
        structure = {
            "root_type": type(data).__name__,
            "item_count": len(data) if hasattr(data, "__len__") else None,
            "preview": data[:20] if isinstance(data, list) else data,
        }
        return data_preview(structure), structure
    if file_type == "pdf":
        reader = PdfReader(str(path))
        text = "\n\n".join(page.extract_text() or "" for page in reader.pages)
        return truncate_text(text), {"page_count": len(reader.pages)}
    raise ValueError(f"Unsupported parser for {file_type}")


async def process_file(db: AsyncSession, uploaded_file: UploadedFile) -> UploadedFile:
    uploaded_file.status = FileStatus.processing
    uploaded_file.error_message = None
    await db.commit()
    try:
        text, data = await asyncio.to_thread(
            _parse_file, Path(uploaded_file.file_path), uploaded_file.file_type
        )
        document_analysis = await analyze_document_async(text or "")
        if isinstance(data, dict):
            data["document_analysis"] = document_analysis
        else:
            data = {"content": data, "document_analysis": document_analysis}
        uploaded_file.extracted_text = text
        uploaded_file.parsed_data = data
        uploaded_file.status = FileStatus.parsed
    except Exception as exc:
        uploaded_file.status = FileStatus.failed
        uploaded_file.error_message = str(exc)[:2000]
    await db.commit()
    await db.refresh(uploaded_file)
    return uploaded_file
